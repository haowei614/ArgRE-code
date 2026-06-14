#!/usr/bin/env python3
"""Generate annotator Excel sheets for the conflict-validation study."""

from __future__ import annotations

import json
import random
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


STUDY_DIR = Path(__file__).resolve().parent
SAMPLES_PATH = STUDY_DIR / "conflict_validation_samples.json"
SHEETS_DIR = STUDY_DIR / "sheets"
ANNOTATOR_SEEDS = {"R1": 7001, "R2": 7002, "R3": 7003}
JUDGMENT_OPTIONS = "Real conflict,Not a real conflict,Borderline"
CONFIDENCE_OPTIONS = "High,Medium,Low"


def main() -> None:
    samples = _load_samples()
    SHEETS_DIR.mkdir(parents=True, exist_ok=True)
    for annotator, seed in ANNOTATOR_SEEDS.items():
        rows = list(samples)
        random.Random(seed).shuffle(rows)
        workbook = Workbook()
        instruction_sheet = workbook.active
        instruction_sheet.title = "Instructions"
        _write_instructions(instruction_sheet)
        annotation_sheet = workbook.create_sheet("Items")
        _write_items(annotation_sheet, rows)
        output_path = SHEETS_DIR / f"conflict_validation_{annotator}.xlsx"
        workbook.save(output_path)
        print(f"Wrote {output_path.relative_to(STUDY_DIR.parent)}")


def _load_samples() -> list[dict[str, Any]]:
    if not SAMPLES_PATH.exists():
        raise FileNotFoundError(
            f"Missing {SAMPLES_PATH}. Run extract_conflict_pairs.py before generating sheets."
        )
    with SAMPLES_PATH.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, list):
        raise ValueError(f"Expected a list in {SAMPLES_PATH}")
    return payload


def _write_instructions(sheet) -> None:
    sheet["A1"] = "Conflict Validation Study Instructions"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A3"] = "Task description"
    sheet["A3"].font = Font(bold=True)
    sheet["A4"] = (
        "Please judge whether each pair of requirements represents a real engineering conflict. "
        "An engineering conflict means the two requirements cannot be fully satisfied "
        "simultaneously in the same system, or satisfying one would significantly compromise "
        "the other."
    )
    sheet["A6"] = "Judgment definitions"
    sheet["A6"].font = Font(bold=True)
    definitions = [
        (
            "Real conflict",
            "The two requirements are incompatible or create a genuine engineering trade-off.",
        ),
        (
            "Not a real conflict",
            "The two requirements can coexist without meaningful tension.",
        ),
        (
            "Borderline",
            "There is potential tension but no clear-cut incompatibility.",
        ),
    ]
    for row_index, (label, definition) in enumerate(definitions, start=7):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=definition)
    sheet["A11"] = "Note"
    sheet["A11"].font = Font(bold=True)
    sheet["A12"] = (
        "Please judge based solely on the requirement text and context provided. "
        "Do not consider how these pairs were selected."
    )
    sheet["A14"] = "Allowed values"
    sheet["A14"].font = Font(bold=True)
    sheet["A15"] = "Judgment"
    sheet["B15"] = JUDGMENT_OPTIONS.replace(",", " / ")
    sheet["A16"] = "Confidence"
    sheet["B16"] = CONFIDENCE_OPTIONS.replace(",", " / ")
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_items(sheet, rows: list[dict[str, Any]]) -> None:
    headers = [
        "Item ID",
        "Case Study",
        "Argument A Text",
        "Argument A Agent",
        "Argument A Round",
        "Argument B Text",
        "Argument B Agent",
        "Argument B Round",
        "Judgment",
        "Confidence",
        "Notes",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for item in rows:
        arg_a = item["argument_a"]
        arg_b = item["argument_b"]
        sheet.append(
            [
                item["item_id"],
                item["case_study"],
                arg_a.get("text", ""),
                arg_a.get("agent", ""),
                arg_a.get("round", ""),
                arg_b.get("text", ""),
                arg_b.get("agent", ""),
                arg_b.get("round", ""),
                "",
                "",
                "",
            ]
        )

    last_row = len(rows) + 1
    judgment_validation = DataValidation(
        type="list", formula1=f'"{JUDGMENT_OPTIONS}"', allow_blank=True
    )
    confidence_validation = DataValidation(
        type="list", formula1=f'"{CONFIDENCE_OPTIONS}"', allow_blank=True
    )
    sheet.add_data_validation(judgment_validation)
    sheet.add_data_validation(confidence_validation)
    judgment_validation.add(f"I2:I{last_row}")
    confidence_validation.add(f"J2:J{last_row}")

    widths = {
        "A": 12,
        "B": 15,
        "C": 80,
        "D": 24,
        "E": 14,
        "F": 80,
        "G": 24,
        "H": 14,
        "I": 24,
        "J": 16,
        "K": 45,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{last_row}"


if __name__ == "__main__":
    main()
