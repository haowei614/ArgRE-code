#!/usr/bin/env python3
"""Analyze completed human conflict-validation annotations."""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STUDY_DIR = Path(__file__).resolve().parent
SHEETS_DIR = STUDY_DIR / "sheets"
RESULTS_DIR = STUDY_DIR / "results"
REVEAL_MAPPING_PATH = STUDY_DIR / "private_reveal_mapping_conflict.json"
RESULTS_PATH = RESULTS_DIR / "conflict_validation_results.json"
ANNOTATOR_FILES = {
    "R1": SHEETS_DIR / "conflict_validation_R1.xlsx",
    "R2": SHEETS_DIR / "conflict_validation_R2.xlsx",
    "R3": SHEETS_DIR / "conflict_validation_R3.xlsx",
}
TERNARY_CATEGORIES = ("Real conflict", "Not a real conflict", "Borderline")
BINARY_CATEGORIES = ("Real conflict", "Not a real conflict")


def main() -> None:
    mapping = _load_mapping()
    annotations = _load_annotations()
    majority = _majority_votes(annotations)
    metrics = _compute_metrics(mapping, annotations, majority)
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    RESULTS_PATH.write_text(json.dumps(metrics, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    _print_results(metrics)


def _load_mapping() -> dict[str, Any]:
    if not REVEAL_MAPPING_PATH.exists():
        raise FileNotFoundError(
            f"Missing {REVEAL_MAPPING_PATH}. Run extract_conflict_pairs.py before analysis."
        )
    with REVEAL_MAPPING_PATH.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict) or "items" not in payload:
        raise ValueError(f"Expected reveal mapping with an 'items' object in {REVEAL_MAPPING_PATH}")
    return payload


def _load_annotations() -> dict[str, dict[str, str]]:
    annotations: dict[str, dict[str, str]] = {}
    for annotator, path in ANNOTATOR_FILES.items():
        if not path.exists():
            print(f"Warning: missing {path}; skipping {annotator}.")
            continue
        workbook = load_workbook(path, data_only=True)
        if "Items" not in workbook.sheetnames:
            print(f"Warning: {path} has no Items sheet; skipping {annotator}.")
            continue
        sheet = workbook["Items"]
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        columns = {name: index for index, name in enumerate(header)}
        if "Item ID" not in columns or "Judgment" not in columns:
            print(f"Warning: {path} is missing Item ID or Judgment columns; skipping {annotator}.")
            continue
        item_col = columns["Item ID"]
        judgment_col = columns["Judgment"]
        annotator_rows: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item_id = row[item_col]
            if not item_id:
                continue
            judgment = _canonical_judgment(row[judgment_col])
            if judgment:
                annotator_rows[str(item_id).strip()] = judgment
        annotations[annotator] = annotator_rows
    return annotations


def _canonical_judgment(value: Any) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    lowered = raw.lower()
    aliases = {
        "real conflict": "Real conflict",
        "conflict": "Real conflict",
        "yes": "Real conflict",
        "y": "Real conflict",
        "not a real conflict": "Not a real conflict",
        "not real conflict": "Not a real conflict",
        "no conflict": "Not a real conflict",
        "no": "Not a real conflict",
        "n": "Not a real conflict",
        "borderline": "Borderline",
        "border line": "Borderline",
        "unclear": "Borderline",
        "maybe": "Borderline",
    }
    return aliases.get(lowered, raw if raw in TERNARY_CATEGORIES else "")


def _majority_votes(annotations: dict[str, dict[str, str]]) -> dict[str, dict[str, Any]]:
    item_ids = sorted({item_id for rows in annotations.values() for item_id in rows})
    majority: dict[str, dict[str, Any]] = {}
    for item_id in item_ids:
        votes = [
            rows[item_id]
            for rows in annotations.values()
            if item_id in rows and rows[item_id] in TERNARY_CATEGORIES
        ]
        counts = Counter(votes)
        threshold = 2 if len(votes) >= 2 else 1
        winner = ""
        if votes:
            label, count = counts.most_common(1)[0]
            tied = sum(1 for value in counts.values() if value == count) > 1
            if count >= threshold and not tied:
                winner = label
        majority[item_id] = {
            "votes": votes,
            "vote_counts": dict(counts),
            "majority": winner,
            "n_votes": len(votes),
        }
    return majority


def _compute_metrics(
    mapping: dict[str, Any],
    annotations: dict[str, dict[str, str]],
    majority: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    items = mapping["items"]
    detected_ids = [item_id for item_id, meta in items.items() if meta["label"] == "detected_attack"]
    negative_ids = [item_id for item_id, meta in items.items() if meta["label"] == "no_attack"]

    precision_overall = _rate(
        item_ids=detected_ids,
        majority=majority,
        positive_label="Real conflict",
    )
    precision_by_pattern = {
        pattern: _rate(
            item_ids=[
                item_id
                for item_id in detected_ids
                if items[item_id].get("pattern") == pattern
            ],
            majority=majority,
            positive_label="Real conflict",
        )
        for pattern in ("P1", "P2", "P3")
    }
    precision_by_case = {
        case: _rate(
            item_ids=[
                item_id
                for item_id in detected_ids
                if items[item_id].get("case_study") == case
            ],
            majority=majority,
            positive_label="Real conflict",
        )
        for case in ("AD", "ATM", "Bookkeeping")
    }
    negative_control_accuracy = _rate(
        item_ids=negative_ids,
        majority=majority,
        positive_label="Not a real conflict",
    )

    agreement = _agreement_metrics(annotations)
    item_vote_summary = {
        item_id: {
            **majority.get(item_id, {"votes": [], "vote_counts": {}, "majority": "", "n_votes": 0}),
            "hidden_label": meta["label"],
            "hidden_pattern": meta.get("pattern"),
            "case_study": meta.get("case_study"),
        }
        for item_id, meta in items.items()
    }

    return {
        "summary": {
            "annotator_count_with_any_judgments": len(
                [rows for rows in annotations.values() if rows]
            ),
            "detected_attack_items": len(detected_ids),
            "negative_control_items": len(negative_ids),
            "items_with_majority_vote": sum(
                1 for item_id in items if majority.get(item_id, {}).get("majority")
            ),
        },
        "precision": precision_overall,
        "precision_by_pattern": precision_by_pattern,
        "precision_by_case_study": precision_by_case,
        "negative_control_accuracy": negative_control_accuracy,
        "agreement": agreement,
        "latex_table": _latex_table(
            precision_overall=precision_overall,
            precision_by_pattern=precision_by_pattern,
            precision_by_case=precision_by_case,
            negative_control_accuracy=negative_control_accuracy,
            agreement=agreement,
        ),
        "item_votes": item_vote_summary,
    }


def _rate(
    *,
    item_ids: list[str],
    majority: dict[str, dict[str, Any]],
    positive_label: str,
) -> dict[str, Any]:
    usable = [
        item_id
        for item_id in item_ids
        if majority.get(item_id, {}).get("majority") in TERNARY_CATEGORIES
    ]
    positive = [
        item_id
        for item_id in usable
        if majority.get(item_id, {}).get("majority") == positive_label
    ]
    value = len(positive) / len(usable) if usable else None
    return {
        "value": value,
        "numerator": len(positive),
        "denominator": len(usable),
        "missing_or_no_majority": len(item_ids) - len(usable),
    }


def _agreement_metrics(annotations: dict[str, dict[str, str]]) -> dict[str, Any]:
    active = [annotator for annotator, rows in annotations.items() if rows]
    if len(active) >= 3:
        complete_ids = sorted(set.intersection(*(set(annotations[a]) for a in active[:3])))
        ternary_ratings = [[annotations[a][item_id] for a in active[:3]] for item_id in complete_ids]
        binary_ratings = [[_binary_vote(vote) for vote in row] for row in ternary_ratings]
        return {
            "method": "Fleiss' kappa",
            "annotators": active[:3],
            "complete_item_count": len(complete_ids),
            "ternary_categories": list(TERNARY_CATEGORIES),
            "ternary_kappa": _fleiss_kappa(ternary_ratings, TERNARY_CATEGORIES),
            "binary_categories": list(BINARY_CATEGORIES),
            "binary_borderline_merged_with_real_conflict_kappa": _fleiss_kappa(
                binary_ratings, BINARY_CATEGORIES
            ),
        }
    if len(active) == 2:
        left, right = active
        complete_ids = sorted(set(annotations[left]).intersection(annotations[right]))
        ternary_pairs = [(annotations[left][item_id], annotations[right][item_id]) for item_id in complete_ids]
        binary_pairs = [(_binary_vote(a), _binary_vote(b)) for a, b in ternary_pairs]
        return {
            "method": "Cohen's kappa",
            "annotators": active,
            "complete_item_count": len(complete_ids),
            "ternary_categories": list(TERNARY_CATEGORIES),
            "ternary_kappa": _cohen_kappa(ternary_pairs, TERNARY_CATEGORIES),
            "binary_categories": list(BINARY_CATEGORIES),
            "binary_borderline_merged_with_real_conflict_kappa": _cohen_kappa(
                binary_pairs, BINARY_CATEGORIES
            ),
        }
    return {
        "method": "insufficient annotators",
        "annotators": active,
        "complete_item_count": 0,
        "ternary_kappa": None,
        "binary_borderline_merged_with_real_conflict_kappa": None,
    }


def _fleiss_kappa(ratings: list[list[str]], categories: tuple[str, ...]) -> float | None:
    if not ratings:
        return None
    n_raters = len(ratings[0])
    if n_raters < 2 or any(len(row) != n_raters for row in ratings):
        return None
    category_totals = Counter()
    p_i_values = []
    for row in ratings:
        counts = Counter(row)
        category_totals.update(counts)
        p_i = (sum(count * count for count in counts.values()) - n_raters) / (
            n_raters * (n_raters - 1)
        )
        p_i_values.append(p_i)
    p_bar = sum(p_i_values) / len(p_i_values)
    total_ratings = len(ratings) * n_raters
    p_e = sum((category_totals[category] / total_ratings) ** 2 for category in categories)
    if p_e == 1:
        return 1.0
    return (p_bar - p_e) / (1 - p_e)


def _cohen_kappa(pairs: list[tuple[str, str]], categories: tuple[str, ...]) -> float | None:
    if not pairs:
        return None
    observed = sum(1 for left, right in pairs if left == right) / len(pairs)
    left_counts = Counter(left for left, _ in pairs)
    right_counts = Counter(right for _, right in pairs)
    expected = sum(
        (left_counts[category] / len(pairs)) * (right_counts[category] / len(pairs))
        for category in categories
    )
    if expected == 1:
        return 1.0
    return (observed - expected) / (1 - expected)


def _binary_vote(vote: str) -> str:
    if vote == "Not a real conflict":
        return "Not a real conflict"
    return "Real conflict"


def _latex_table(
    *,
    precision_overall: dict[str, Any],
    precision_by_pattern: dict[str, dict[str, Any]],
    precision_by_case: dict[str, dict[str, Any]],
    negative_control_accuracy: dict[str, Any],
    agreement: dict[str, Any],
) -> str:
    rows = [
        ("Detected attacks precision", precision_overall),
        ("P1 precision", precision_by_pattern["P1"]),
        ("P2 precision", precision_by_pattern["P2"]),
        ("P3 precision", precision_by_pattern["P3"]),
        ("AD precision", precision_by_case["AD"]),
        ("ATM precision", precision_by_case["ATM"]),
        ("Bookkeeping precision", precision_by_case["Bookkeeping"]),
        ("Negative-control accuracy", negative_control_accuracy),
    ]
    table_lines = [
        r"\begin{tabular}{lcc}",
        r"\toprule",
        r"Metric & Value & n \\",
        r"\midrule",
    ]
    for label, metric in rows:
        table_lines.append(
            f"{label} & {_format_value(metric['value'])} & "
            f"{metric['numerator']}/{metric['denominator']} \\\\"
        )
    table_lines.extend(
        [
            r"\midrule",
            (
                f"{agreement['method']} (3-category) & "
                f"{_format_value(agreement.get('ternary_kappa'))} & "
                f"{agreement.get('complete_item_count', 0)} \\\\"
            ),
            (
                f"{agreement['method']} (binary) & "
                f"{_format_value(agreement.get('binary_borderline_merged_with_real_conflict_kappa'))} & "
                f"{agreement.get('complete_item_count', 0)} \\\\"
            ),
            r"\bottomrule",
            r"\end{tabular}",
        ]
    )
    return "\n".join(table_lines)


def _print_results(metrics: dict[str, Any]) -> None:
    print("Conflict-validation analysis summary")
    print(json.dumps(metrics["summary"], indent=2, ensure_ascii=False))
    print("\nPrecision:")
    _print_metric("Overall detected attacks", metrics["precision"])
    for pattern, metric in metrics["precision_by_pattern"].items():
        _print_metric(pattern, metric)
    print("\nPrecision by case study:")
    for case, metric in metrics["precision_by_case_study"].items():
        _print_metric(case, metric)
    print("\nNegative controls:")
    _print_metric("Negative-control accuracy", metrics["negative_control_accuracy"])
    print("\nAgreement:")
    agreement = metrics["agreement"]
    print(f"  Method: {agreement['method']}")
    print(f"  Complete items: {agreement.get('complete_item_count', 0)}")
    print(f"  Ternary kappa: {_format_value(agreement.get('ternary_kappa'))}")
    print(
        "  Binary kappa (Borderline merged with Real conflict): "
        f"{_format_value(agreement.get('binary_borderline_merged_with_real_conflict_kappa'))}"
    )
    print("\nLaTeX table:")
    print(metrics["latex_table"])
    print(f"\nWrote {RESULTS_PATH.relative_to(STUDY_DIR.parent)}")


def _print_metric(label: str, metric: dict[str, Any]) -> None:
    print(
        f"  {label}: {_format_value(metric['value'])} "
        f"({metric['numerator']}/{metric['denominator']}; "
        f"missing/no majority={metric['missing_or_no_majority']})"
    )


def _format_value(value: Any) -> str:
    if value is None:
        return "N/A"
    return f"{float(value):.3f}"


if __name__ == "__main__":
    main()
